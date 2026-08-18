#!/usr/bin/env python3
"""Consolidated DAM sleep-analysis pipeline.

Combines, in order:
  0) (Optional) Import raw multi-channel Monitor*.txt dumps -- straight off the
     DAM hardware, one file per monitor, all channels and measurement types
     mixed together -- into the legacy per-channel "Raw Data" format Step 1
     expects, using a plain-text channel-groups manifest (monitor, group,
     channel range) instead of an Excel channelList. Skip this step if you
     already have a "Raw Data" folder of one-file-per-channel .txt files.
  1) Per-experiment channel processing + activity/sleep binning + PDF plots
     (previously basicDAManalysis_v3_mac.m)
  2) Multi-experiment aggregation into a GraphPad-ready multi-sheet Excel file
     (previously mat2write_dayVsNightSleepBins_graphPadFormat_MAC.m)
  3) Removal of dead/incomplete animals from that Excel file
     (previously clean_dead_MAC.m)
  4) Prism-ready plots (bar+dot, sex-split scatter, ZT sleep profile, and a
     continuous multi-day sleep profile) + copy/paste-ready data tables
     (sleep_pipeline.py, imported directly -- must sit next to this file)

Run with:  python3 run_sleep_analysis_pipeline.py
You will be prompted (via native folder/file dialogs if a display is
available, otherwise via the terminal) instead of having to hand-edit
variables at the top of a script. Uses only pathlib for file paths, so it
runs unmodified on macOS and Windows.

Dependencies: numpy, matplotlib, openpyxl, python-dateutil (tkinter is used
for dialogs when available, but the script falls back to the terminal if
tkinter or a display is not available, e.g. over SSH).

Per-experiment intermediate data is stored as a pickle file (<name>.pkl)
next to the channelList Excel file -- this pipeline does not read/write
MATLAB .mat files, so it is fully self-contained in Python.
"""

import math
import pickle
import re
import sys
from datetime import timedelta
from pathlib import Path

import numpy as np
from dateutil import parser as dateparser
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib
import matplotlib.pyplot as plt
import openpyxl

matplotlib.use("Agg")

try:
    import tkinter as tk
    from tkinter import filedialog, simpledialog

    _root = tk.Tk()
    _root.withdraw()
    HAS_GUI = True
except Exception:
    HAS_GUI = False


# =============================================================================
# Prompt helpers (dialogs when a display is available, terminal otherwise)
# =============================================================================
def ask_folder(prompt_text, default_path=None, create_if_missing=False):
    default_path = str(default_path) if default_path else str(Path.cwd())
    if HAS_GUI:
        path = filedialog.askdirectory(title=prompt_text, initialdir=default_path)
        if not path:
            raise SystemExit(f"{prompt_text}: no folder selected.")
    else:
        txt = input(f"{prompt_text} [{default_path}]: ").strip()
        path = txt or default_path
    path = Path(path)
    if not path.is_dir():
        if create_if_missing:
            path.mkdir(parents=True, exist_ok=True)
        else:
            raise SystemExit(f"Folder does not exist: {path}")
    return path


def ask_file(prompt_text, default_folder=None, filetypes=(("Excel files", "*.xlsx"),)):
    default_folder = str(default_folder) if default_folder else str(Path.cwd())
    if HAS_GUI:
        path = filedialog.askopenfilename(title=prompt_text, initialdir=default_folder, filetypes=filetypes)
        if not path:
            raise SystemExit(f"{prompt_text}: no file selected.")
    else:
        txt = input(f"{prompt_text} (full path): ").strip()
        if not txt:
            raise SystemExit(f"{prompt_text}: no file entered.")
        path = txt
    path = Path(path)
    if not path.is_file():
        raise SystemExit(f"File does not exist: {path}")
    return path


def ask_number(prompt_text, default_val):
    if HAS_GUI:
        val = simpledialog.askinteger("Input", prompt_text, initialvalue=default_val)
        if val is None:
            raise SystemExit(f"{prompt_text}: no value entered.")
        return val
    txt = input(f"{prompt_text} [{default_val}]: ").strip()
    return int(txt) if txt else default_val


def ask_choice(prompt_text, options, default_idx=0):
    if HAS_GUI:
        txt = simpledialog.askstring(
            "Input",
            f"{prompt_text}\nOptions: " + ", ".join(f"{i+1}={o}" for i, o in enumerate(options)),
            initialvalue=str(default_idx + 1),
        )
        idx = int(txt) - 1 if txt and txt.strip() else default_idx
    else:
        print(prompt_text)
        for i, o in enumerate(options):
            print(f"  {i+1}) {o}")
        txt = input(f"Choice [{default_idx+1}]: ").strip()
        idx = int(txt) - 1 if txt else default_idx
    return options[idx]


ZT_RANGE_RE = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)\s*$")


def fmt_zt_hour(h):
    """0/24/12 -> '0'/'24'/'12'; 3.5 -> '3p5' (dot isn't filename-safe)."""
    return str(int(h)) if h == int(h) else str(h).replace(".", "p")


def ask_zt_windows(default="0-24"):
    prompt_text = (
        "Which ZT hour range(s) do you want to analyze? Each range is 'start-end' within 0-24, "
        "in 0.5h steps (the data is 30-min binned) -- e.g. '0-24' for the full day, '0-12, 12-24' "
        "for day vs night, or '0-3' for just the first 3 hours. Comma-separate multiple ranges "
        "to get one output file per range."
    )
    if HAS_GUI:
        txt = simpledialog.askstring("Input", prompt_text, initialvalue=default)
    else:
        print(prompt_text)
        txt = input(f"ZT range(s) [{default}]: ").strip()
    txt = (txt or default).strip()

    windows = []
    for part in txt.split(","):
        part = part.strip()
        if not part:
            continue
        m = ZT_RANGE_RE.match(part)
        if not m:
            raise SystemExit(f"Could not parse ZT range '{part}'. Use the form 'start-end', e.g. '0-24'.")
        lo, hi = float(m.group(1)), float(m.group(2))
        if not (0 <= lo < hi <= 24):
            raise SystemExit(f"ZT range '{part}' must satisfy 0 <= start < end <= 24.")
        if (lo * 2) % 1 != 0 or (hi * 2) % 1 != 0:
            raise SystemExit(f"ZT range '{part}': start/end must be in 0.5-hour steps (data is 30-min binned).")
        windows.append((lo, hi))
    if not windows:
        raise SystemExit("No ZT range entered.")
    return windows


def ask_datetime_range():
    prompt_start = "Start date & time of the range to analyze, e.g. '2026-07-28 10:10:00'"
    prompt_end = ("End date & time of the range to analyze -- the LAST minute you want included "
                  "(same convention as DAMFileScan), e.g. '2026-07-30 10:09:00'")
    if HAS_GUI:
        start_txt = simpledialog.askstring("Input", prompt_start)
        end_txt = simpledialog.askstring("Input", prompt_end)
    else:
        start_txt = input(f"{prompt_start}: ").strip()
        end_txt = input(f"{prompt_end}: ").strip()

    try:
        start_dt = dateparser.parse(start_txt)
        end_dt = dateparser.parse(end_txt)
    except (ValueError, OverflowError, TypeError):
        raise SystemExit(f"Could not parse '{start_txt}' / '{end_txt}' as dates.")
    if not start_dt or not end_dt or end_dt < start_dt:
        raise SystemExit("End date/time must be at or after start date/time.")

    # end_dt is the last included minute (DAMFileScan convention), so the count is inclusive.
    total_minutes = int(round((end_dt - start_dt).total_seconds() / 60)) + 1
    n_days = total_minutes // 1440
    if n_days < 1:
        raise SystemExit(f"Range is only {total_minutes} minute(s) -- must span at least one full day (1440 minutes).")
    if total_minutes % 1440 != 0:
        print(f"Note: range is {total_minutes} minutes, not a whole number of days -- "
              f"truncating to {n_days} whole day(s) ({n_days * 1440} minutes) starting {start_dt}.")
    end_dt = start_dt + timedelta(minutes=n_days * 1440)
    return start_dt, end_dt, n_days


# =============================================================================
# Step 0: import raw Monitor*.txt dumps into the legacy per-channel format
# =============================================================================
def parse_monitor_dump(path):
    """Parse a raw multi-channel TriKinetics Monitor*.txt dump. Each line is one
    minute of ALL channels for one measurement type: col2=date, col3=time,
    col8=type (CT/MT/Pn), col11 onward = one activity count per channel.
    Returns dict: measurement_type -> {datetime: np.ndarray of per-channel counts}."""
    by_type = {}
    with open(path, "r", errors="replace") as f:
        for line in f:
            fields = line.rstrip("\n").split("\t")
            if len(fields) < 11:
                fields = line.split()
            if len(fields) < 11:
                continue
            try:
                dt = dateparser.parse(f"{fields[1]} {fields[2]}")
            except (ValueError, OverflowError, TypeError):
                continue
            mtype = fields[7].strip().upper()  # normalize casing (real files have e.g. both 'CT' and 'Ct')
            try:
                counts = np.array([float(c) for c in fields[10:]], dtype=float)
            except ValueError:
                continue
            by_type.setdefault(mtype, {})[dt] = counts
    return by_type


def parse_channel_range(range_str):
    """'1-16' or '1,3,5-8' (';' also accepted as a separator) -> sorted channel list."""
    channels = []
    for part in range_str.replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            lo, hi = part.split("-")
            channels.extend(range(int(lo), int(hi) + 1))
        else:
            channels.append(int(part))
    return sorted(set(channels))


def read_channel_groups_manifest(path):
    """Plain-text manifest, one line per monitor+group+channel-range block:
        monitor, group, channel_range
        38, 23E10GS_pex16_male, 1-16
        38, 23E10GS_pex16_female, 17-32
    '#' starts a comment; blank lines are ignored. Returns a list of
    (channel, group, monitor) tuples, one per channel, sorted by monitor then channel."""
    expanded = []
    seen = set()
    with open(path, "r") as f:
        for lineno, raw_line in enumerate(f, start=1):
            line = raw_line.split("#", 1)[0].strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]
            if len(parts) != 3:
                raise SystemExit(f"{path}:{lineno}: expected 'monitor, group, channel_range', got: {raw_line.strip()!r}")
            monitor = int(parts[0])
            group = parts[1]
            for ch in parse_channel_range(parts[2]):
                if (monitor, ch) in seen:
                    raise SystemExit(f"{path}:{lineno}: monitor {monitor} channel {ch} is listed more than once.")
                seen.add((monitor, ch))
                expanded.append((ch, group, monitor))
    if not expanded:
        raise SystemExit(f"{path}: no channel entries found.")
    expanded.sort(key=lambda t: (t[2], t[0]))
    return expanded


def choose_measurement_type(types_by_monitor):
    all_types = set()
    for s in types_by_monitor.values():
        all_types |= s
    if len(all_types) == 1:
        return next(iter(all_types))
    options = sorted(all_types)
    default_idx = options.index("CT") if "CT" in options else 0
    return ask_choice(
        "Multiple measurement types found across the monitor files (e.g. CT/MT/Pn). "
        "Which should be used as activity counts? (CT -- beam-crossing counts -- is the standard choice.)",
        options, default_idx,
    )


def step0_import_monitor_data(state):
    print("\n--- Step 0: Import raw Monitor data ---")
    dump_folder = ask_folder("Select the folder containing the raw Monitor*.txt files", state.get("last_folder"))
    dest_folder = ask_folder(
        "Select (or create) the destination experiment folder -- this is where 'Raw Data' and the "
        "channelList will be written, and what you'll point Step 1 at",
        state.get("last_folder"),
        create_if_missing=True,
    )

    # The manifest describes THIS experiment's monitor/group/channel layout, so it
    # belongs with the experiment's other files, not the (often-reused) raw dump
    # folder -- same naming convention as channelList: "<expName>_channelGroups.txt".
    manifest_path = dest_folder / f"{dest_folder.name}_channelGroups.txt"
    if not manifest_path.is_file():
        manifest_path = ask_file(
            "Select the channel-groups manifest (plain text: monitor, group, channel_range per line)",
            dest_folder,
            filetypes=(("Text files", "*.txt"),),
        )

    start_dt, end_dt, n_days = ask_datetime_range()

    manifest = read_channel_groups_manifest(manifest_path)
    monitors_needed = sorted({m for _, _, m in manifest})

    print(f"Loading {len(monitors_needed)} monitor file(s): {monitors_needed}")
    parsed_by_monitor = {}
    types_by_monitor = {}
    for mon in monitors_needed:
        mon_path = dump_folder / f"Monitor{mon}.txt"
        if not mon_path.is_file():
            raise SystemExit(f"Could not find {mon_path}")
        by_type = parse_monitor_dump(mon_path)
        if not by_type:
            raise SystemExit(f"No parsable rows found in {mon_path}")
        parsed_by_monitor[mon] = by_type
        types_by_monitor[mon] = set(by_type.keys())
        print(f"  Monitor {mon}: types found = {sorted(by_type.keys())}")

    mtype = choose_measurement_type(types_by_monitor)
    print(f"Using measurement type: {mtype}")

    raw_dir = dest_folder / "Raw Data"
    raw_dir.mkdir(exist_ok=True)

    expected_minutes = [start_dt + timedelta(minutes=i) for i in range(n_days * 1440)]
    date_tag = start_dt.strftime("%Y%m%d")
    total_missing = 0

    for channel, group, monitor in manifest:
        if mtype not in parsed_by_monitor[monitor]:
            raise SystemExit(
                f"Monitor {monitor} has no '{mtype}' rows (found: {sorted(types_by_monitor[monitor])})."
            )
        series_by_dt = parsed_by_monitor[monitor][mtype]

        counts = np.full(len(expected_minutes), np.nan)
        missing = 0
        for i, dt in enumerate(expected_minutes):
            row = series_by_dt.get(dt)
            if row is None:
                missing += 1
                continue
            if channel - 1 >= len(row):
                raise SystemExit(
                    f"Monitor {monitor} row at {dt} has only {len(row)} channel(s); channel {channel} was requested."
                )
            counts[i] = row[channel - 1]
        total_missing += missing
        if missing:
            print(f"  Monitor {monitor} channel {channel}: {missing} minute(s) missing from the raw file (filled with NaN).")

        out_path = raw_dir / f"{date_tag}M{monitor:03d}C{channel:02d}.txt"
        with open(out_path, "w") as f:
            f.write(f"Monitor{monitor}C{channel:02d}   {start_dt.strftime('%d %b %y')} {start_dt.strftime('%H:%M:%S')}\n")
            f.write(f"{len(expected_minutes)}\n")
            f.write("0\n0\n")
            for v in counts:
                f.write(("nan" if math.isnan(v) else str(int(v))) + "\n")

    exp_name = dest_folder.name
    channel_list_path = dest_folder / f"{exp_name}_channelList.txt"
    with open(channel_list_path, "w") as f:
        f.write("# channel, group, monitor -- auto-generated by Step 0 from the channel-groups manifest\n")
        for channel, group, monitor in manifest:
            f.write(f"{channel},{group},{monitor}\n")

    # Also auto-generate the mat2read entry for this one experiment, so Step 2 can
    # find it without you having to hand-build a second file -- it's just this
    # folder + this channelList + the group names from your manifest. If you later
    # want to pool in OTHER experiments too, edit this file and add more rows.
    mat2read_path = dest_folder / f"{exp_name}_mat2read.txt"
    if not mat2read_path.is_file():
        groups_in_order = list(dict.fromkeys(g for _, g, _ in manifest))
        with open(mat2read_path, "w") as f:
            f.write("# rootdir, channelList base name, group columns...\n")
            f.write(",".join([str(dest_folder), f"{exp_name}_channelList", *groups_in_order]) + "\n")
        print(f"Wrote {mat2read_path} (edit it to add more experiments before running Step 2)")

    print(f"Wrote {len(manifest)} channel file(s) to {raw_dir}")
    print(f"Wrote {channel_list_path}")
    if total_missing:
        print(f"Note: {total_missing} total minute(s) were missing across all channels and filled with NaN.")

    state["last_folder"] = dest_folder
    state["max_days"] = n_days
    return state


# =============================================================================
# Step 1: per-experiment DAM processing
# =============================================================================
def find_true_runs(mask):
    """Return list of (start, end) inclusive 0-based indices of True runs."""
    runs = []
    n = len(mask)
    i = 0
    while i < n:
        if mask[i]:
            j = i
            while j + 1 < n and mask[j + 1]:
                j += 1
            runs.append((i, j))
            i = j + 1
        else:
            i += 1
    return runs


DATE_IN_LINE_RE = re.compile(r"\d{1,2}\s+[A-Za-z]{3,9}\.?\s+\d{2,4}")


def extract_date_from_line(line):
    """DAM header lines look like '<channel-id><spaces><DD Mon YYYY>[ HH:MM:SS]'.
    Search for the date pattern rather than fuzzy-parsing the whole line --
    the leading channel/monitor id (all digits) confuses dateutil's fuzzy
    matcher into raising instead of skipping it."""
    match = DATE_IN_LINE_RE.search(line)
    if match:
        return dateparser.parse(match.group(0))
    return dateparser.parse(line, fuzzy=True)


def parse_channel_file_multi_file_per_day(path):
    """One DAM .txt file = one day. Returns (day_date, counts array)."""
    with open(path, "r", errors="replace") as f:
        lines = f.readlines()
    day_date = extract_date_from_line(lines[0])

    num_sample_pts = int(float(lines[1].strip()))
    if num_sample_pts <= 1440:
        num_sample_pts = 1440
    # lines[2], lines[3] are discarded (headers)
    counts = np.full(num_sample_pts, np.nan)
    data_lines = lines[4:4 + num_sample_pts]
    for i, line in enumerate(data_lines):
        counts[i] = float(line.strip())
    return day_date, counts


def parse_channel_file_multi_day_single_file(path, max_days_hint=None):
    """A single DAM .txt file spans multiple days (already split into 1440-sample chunks)."""
    with open(path, "r", errors="replace") as f:
        lines = f.readlines()
    day_date = extract_date_from_line(lines[0])
    num_sample_pts = int(float(lines[1].strip()))
    data_lines = lines[4:4 + num_sample_pts]
    counts = np.full(num_sample_pts, np.nan)
    for i, line in enumerate(data_lines):
        counts[i] = float(line.strip())

    days = []
    for start in range(0, num_sample_pts - num_sample_pts % 1440, 1440):
        days.append(counts[start:start + 1440])
    return day_date, days


def process_channel(raw_dir, channel_num, monitor_num=None):
    """Read all DAM files for one channel and return a list of per-day arrays, sorted chronologically."""
    if monitor_num is not None:
        pattern = f"*M{int(monitor_num):03d}C{int(channel_num):02d}.txt"
    else:
        pattern = f"*C{int(channel_num):02d}.txt"
    matches = sorted(Path(raw_dir).glob(pattern))

    if len(matches) > 1:
        dated_days = [parse_channel_file_multi_file_per_day(p) for p in matches]
        dated_days.sort(key=lambda x: x[0])
        return [counts for _, counts in dated_days]
    elif len(matches) == 1:
        _, days = parse_channel_file_multi_day_single_file(matches[0])
        return days
    else:
        raise FileNotFoundError(f"Could not find files matching {pattern} in {raw_dir}")


def bin_30min(counts_1440):
    """Sum 1-min samples into 48 half-hour bins (matches MATLAB reshape(v,30,48) column sums)."""
    return counts_1440.reshape(48, 30).sum(axis=1)


def compute_sleep_binary(counts_1440):
    """A stop of >=5 consecutive 1-min bins of zero activity counts as sleep."""
    is_stopped = counts_1440 == 0
    is_sleep = np.zeros_like(counts_1440)
    for start, end in find_true_runs(is_stopped):
        if (end - start + 1) >= 5:
            is_sleep[start:end + 1] = 1
    return is_sleep


def detect_behavioral_death(continuous_counts, window_min=1440, prop_immobile=0.01, step_min=60):
    """Behavioral (movement-based) death detection, matching the logic of rethomics/
    sleepr's curate_dead_animals(): scan FORWARD through the fly's whole 1-min activity
    time series (all days concatenated) looking for the first window of `window_min`
    consecutive minutes where the fraction of "moving" minutes (count > 0) is below
    `prop_immobile`. A new window is checked every `step_min` minutes (rethomics calls
    this "resolution"; e.g. a 1440-min window checked every 60 min matches rethomics'
    default of resolution=24, i.e. window_min/24).

    NaN minutes (missing/malformed data) inside a window are excluded from the moving
    fraction rather than counted as immobile, so a data gap alone can't masquerade as
    death here -- that's the separate, existing gap-based check.

    Returns the 0-based minute index where the first qualifying window STARTS (i.e. the
    estimated onset of sustained immobility), or None if no such window is found. Only
    windows with a full window_min of *future* data available are checked, same as
    rethomics scanning "the right (future) data" -- so a fly can't be flagged as dead
    in its last window_min-1 minutes just because there isn't enough data left to check.
    """
    n = len(continuous_counts)
    if n < window_min:
        return None
    moving = continuous_counts > 0
    valid = ~np.isnan(continuous_counts)
    for start in range(0, n - window_min + 1, step_min):
        window_valid = valid[start:start + window_min]
        if not window_valid.any():
            continue
        frac_moving = moving[start:start + window_min][window_valid].mean()
        if frac_moving < prop_immobile:
            return start
    return None


def read_channel_list_txt(path):
    """channelList as plain text (e.g. produced by Step 0): 'channel,group,monitor' per
    line, no header. '#' starts a comment; monitor is optional."""
    rows = []
    with open(path, "r") as f:
        for line in f:
            line = line.split("#", 1)[0].strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]
            if len(parts) < 2:
                continue
            channel = int(parts[0])
            group = parts[1]
            monitor = int(parts[2]) if len(parts) > 2 and parts[2] else None
            rows.append((channel, group, monitor))
    return rows


def step1_process_experiment(state):
    print("\n--- Step 1: Process one experiment ---")
    rootdir = ask_folder(
        'Select the experiment folder (contains "Raw Data" and the channelList file)',
        state.get("last_folder"),
    )
    state["last_folder"] = rootdir
    exp_name = rootdir.name

    raw_dir = rootdir / "Raw Data"
    if not raw_dir.is_dir():
        raise SystemExit(f'Could not find a "Raw Data" folder inside {rootdir}')

    channel_list_file = rootdir / f"{exp_name}_channelList.xlsx"
    if not channel_list_file.is_file():
        channel_list_file = rootdir / f"{exp_name}_channelList.txt"
    if not channel_list_file.is_file():
        channel_list_file = ask_file(
            "Select the channelList file (.xlsx or .txt)",
            rootdir,
            filetypes=(("channelList files", "*.xlsx *.txt"), ("Excel files", "*.xlsx"), ("Text files", "*.txt")),
        )

    default_max_days = state.get("max_days") or 3
    max_days = ask_number("How many days were recorded?", default_max_days)
    state["max_days"] = max_days

    detect_death = ask_choice(
        "Detect flies with behavioral death (sustained immobility) and fully exclude "
        "them from the export -- separate from, and in addition to, flies later removed "
        "in Step 3 for missing/incomplete data?",
        ["Yes", "No"], 0,
    ) == "Yes"
    if detect_death:
        death_window_hours = ask_number(
            "Behavioral-death window (hours) -- a fly moving less than the threshold "
            "below, continuously, for this long counts as dead (rethomics/sleepr default: 24)",
            24,
        )
        death_prop_immobile_pct = ask_number(
            "Max percent of that window the fly can be moving and still be called dead "
            "(rethomics/sleepr default: 1)",
            1,
        )
        death_window_min = death_window_hours * 60
        death_prop_immobile = death_prop_immobile_pct / 100.0
        death_step_min = max(1, death_window_min // 24)  # matches rethomics' default resolution=24
    else:
        death_window_min = death_prop_immobile = death_step_min = None

    if channel_list_file.suffix.lower() == ".txt":
        rows = read_channel_list_txt(channel_list_file)
    else:
        wb = openpyxl.load_workbook(channel_list_file, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    channel_nums = [r[0] for r in rows]
    group_names_by_channel = [r[1] for r in rows]
    monitor_nums = [r[2] if len(r) > 2 else None for r in rows]
    have_monitor_info = all(m is not None for m in monitor_nums)

    group_names = list(dict.fromkeys(g for g in group_names_by_channel if g))
    cmap = plt.get_cmap("tab10")
    group_colors = {g: cmap(i % 10) for i, g in enumerate(group_names)}

    output_base = channel_list_file.with_suffix("")
    all_channel_dat_by_day = []
    for idx, group in enumerate(group_names_by_channel):
        ch = channel_nums[idx]
        mon = monitor_nums[idx] if have_monitor_info else None
        try:
            all_channel_dat_by_day.append(process_channel(raw_dir, ch, mon))
        except FileNotFoundError as exc:
            print(f"Warning: {exc}")
            all_channel_dat_by_day.append([])

    with open(f"{output_base}.pkl", "wb") as f:
        pickle.dump({"all_channel_dat_by_day": all_channel_dat_by_day}, f)
    print(f"Saved {output_base}.pkl")

    pdf_path = f"{output_base}.pdf"
    all_dat_by_group = {}
    excluded_flies = []
    with PdfPages(pdf_path) as pdf:
        for group in group_names:
            print(f"Group: {group}")
            group_activity = {d: None for d in range(1, max_days + 1)}
            sleep_runs_by_fly = []
            active_runs_by_fly = []
            for idx, g in enumerate(group_names_by_channel):
                if g != group:
                    continue
                channel_days = all_channel_dat_by_day[idx]
                if not channel_days:
                    continue

                # Build the continuous (all days concatenated) counts array for this fly
                # BEFORE any sleep classification, so a bout spanning midnight isn't
                # artificially fragmented at the day boundary. A malformed/missing day
                # is replaced with NaN (not skipped) so every fly's day indices stay
                # aligned to calendar days.
                capped_days = list(channel_days[:max_days])
                for i, counts in enumerate(capped_days):
                    if counts.size != 1440:
                        print(f"Channel row {idx} day {i+1}: expected 1440 samples, got {counts.size} - "
                              f"treating as missing (NaN) to keep day alignment.")
                        capped_days[i] = np.full(1440, np.nan)
                if len(capped_days) < max_days:
                    capped_days.extend([np.full(1440, np.nan)] * (max_days - len(capped_days)))
                continuous_counts = np.concatenate(capped_days)

                # --- Behavioral death check: runs on the raw 1-min data, BEFORE any
                # summarizing, so it catches a fly at the earliest possible point and
                # results in the fly being fully excluded from every day of the numeric
                # export -- not just trimmed after the fact. The fly is still plotted
                # below so you can see what happened to it in the PDF. Gap/missing-data
                # exclusion is handled separately, downstream, in Step 3.
                death_idx = None
                if detect_death:
                    death_idx = detect_behavioral_death(
                        continuous_counts, death_window_min, death_prop_immobile, death_step_min
                    )

                excluded = death_idx is not None
                if excluded:
                    death_day = death_idx // 1440 + 1
                    death_zt = (death_idx % 1440) / 60.0
                    reason_text = f"behavioral death (day {death_day} ZT {death_zt:.1f})"
                    mon_label = monitor_nums[idx] if have_monitor_info else None
                    print(f"  Channel {channel_nums[idx]} ({group}): excluded -- {reason_text}. "
                          f"Still plotted in the PDF for reference.")
                    excluded_flies.append({
                        "channel": channel_nums[idx], "monitor": mon_label, "group": group,
                        "reason": reason_text,
                    })

                continuous_is_sleep = compute_sleep_binary(continuous_counts)
                if not excluded:
                    sleep_runs_by_fly.append(find_true_runs(continuous_is_sleep == 1))
                    active_runs_by_fly.append(find_true_runs(continuous_is_sleep == 0))

                fig, axes = plt.subplots(max_days, 2, figsize=(8.5, 11))
                axes = np.atleast_2d(axes)
                for di in range(1, max_days + 1):
                    counts = capped_days[di - 1]
                    is_sleep = continuous_is_sleep[(di - 1) * 1440: di * 1440]
                    activity30 = bin_30min(counts)
                    timepts = np.arange(len(activity30)) / 2.0

                    ax_act = axes[di - 1, 0]
                    ax_act.plot(timepts, activity30, color=group_colors[group])
                    ax_act.set_xlim(0, 24)

                    sleep30 = bin_30min(is_sleep)
                    mins_awake = 1440 - is_sleep.sum()

                    ax_sleep = axes[di - 1, 1]
                    ax_sleep.plot(timepts, sleep30, color=group_colors[group])
                    ax_sleep.set_xlim(0, 24)
                    ax_sleep.set_ylim(0, 30)

                    day_entry = group_activity[di]
                    if not excluded:
                        if day_entry is None:
                            group_activity[di] = {
                                "activity30": activity30[None, :],
                                "activity1": counts[None, :],
                                "sleep30": sleep30[None, :],
                                "mins_awake": np.array([mins_awake]),
                                "sleep_binary": is_sleep[None, :],
                            }
                        else:
                            day_entry["activity30"] = np.vstack([day_entry["activity30"], activity30])
                            day_entry["activity1"] = np.vstack([day_entry["activity1"], counts])
                            day_entry["sleep30"] = np.vstack([day_entry["sleep30"], sleep30])
                            day_entry["mins_awake"] = np.append(day_entry["mins_awake"], mins_awake)
                            day_entry["sleep_binary"] = np.vstack([day_entry["sleep_binary"], is_sleep])

                axes[0, 0].set_ylabel("Activity / 30 min bin")
                mon_label = monitor_nums[idx] if have_monitor_info else "?"
                title = f"{exp_name} M{mon_label} Ch{channel_nums[idx]}: {group}"
                if excluded:
                    axes[0, 0].set_title(f"{title}\nEXCLUDED -- {reason_text}", color="firebrick")
                else:
                    axes[0, 0].set_title(title)
                axes[0, 1].set_ylabel("Sleep (mins)")
                fig.tight_layout()
                pdf.savefig(fig)
                plt.close(fig)
            group_activity["sleep_runs"] = sleep_runs_by_fly
            group_activity["active_runs"] = active_runs_by_fly
            all_dat_by_group[group] = group_activity

    with open(f"{output_base}.pkl", "wb") as f:
        pickle.dump(
            {"all_channel_dat_by_day": all_channel_dat_by_day, "all_dat_by_group": all_dat_by_group},
            f,
        )
    print(f"Saved {output_base}.pkl (with group summaries) and {pdf_path}")

    if excluded_flies:
        report_path = f"{output_base}_behavioral_dead_flies.csv"
        with open(report_path, "w") as f:
            f.write("channel,monitor,group,reason\n")
            for row in excluded_flies:
                f.write(f"{row['channel']},{row['monitor']},{row['group']},\"{row['reason']}\"\n")
        print(f"Excluded {len(excluded_flies)} fly(ies) for behavioral death (sustained immobility). "
              f"Details in {report_path}. All channels, including excluded ones, are still in the PDF. "
              f"Flies with missing/incomplete data are handled separately in Step 3.")

    return state


# =============================================================================
# Step 2: multi-experiment GraphPad export
# =============================================================================
def gather_group_day_data(raw_rows, group_col_idx, day, zt_lo, zt_hi):
    """group_col_idx is 0-based index into the group columns (0 = first group column).
    zt_lo/zt_hi are the ZT hour window (0-24, 0.5h steps) to aggregate over.
    Returns a dict of per-fly arrays/lists (rows aligned across all of them), or
    None if no data was found for this group/day."""
    activity_per_min = []
    sleep_per_min = []
    mins_awake = []
    sleep_binary = []
    activity1_zt = []
    exp_num = []
    sleep_runs_all = []
    active_runs_all = []
    group_name_2_match = None

    zt_lo_min, zt_hi_min = int(round(zt_lo * 60)), int(round(zt_hi * 60))
    zt_lo_bin, zt_hi_bin = int(round(zt_lo * 2)), int(round(zt_hi * 2))
    day_global_offset = (day - 1) * 1440
    win_lo = day_global_offset + zt_lo_min
    win_hi = day_global_offset + zt_hi_min

    for ri, row in enumerate(raw_rows, start=1):
        rootdir, matname = row[0], row[1]
        group_name_2_match = row[2 + group_col_idx]

        if rootdir is None or (isinstance(rootdir, float) and math.isnan(rootdir)):
            print(f"Row {ri}: no rootdir listed - skipping.")
            continue

        mat_path = Path(rootdir) / str(matname)
        if not mat_path.is_file():
            mat_path = Path(rootdir) / f"{matname}.pkl"
        if not mat_path.is_file():
            print(f"Could not find {matname}(.pkl) for row {ri} - skipping.")
            continue

        with open(mat_path, "rb") as f:
            M = pickle.load(f)
        group_dat_for_exp = M.get("all_dat_by_group", {})
        if group_name_2_match not in group_dat_for_exp:
            print(f'Group "{group_name_2_match}" not found in {mat_path} - skipping row {ri}.')
            continue

        group_dat = group_dat_for_exp[group_name_2_match]
        day_entry = group_dat.get(day)
        if day_entry is None:
            continue

        sleep_bin_full = day_entry["sleep_binary"]
        activity_full = day_entry["activity30"]
        activity1_full = day_entry["activity1"]
        sleep_bin_zt = sleep_bin_full[:, zt_lo_min:zt_hi_min]
        this_mins_awake = sleep_bin_zt.shape[1] - np.nansum(sleep_bin_zt, axis=1)
        n_flies_this_exp = sleep_bin_full.shape[0]

        activity_per_min.append(activity_full[:, zt_lo_bin:zt_hi_bin])
        sleep_per_min.append(day_entry["sleep30"])
        mins_awake.append(this_mins_awake)
        sleep_binary.append(sleep_bin_zt)
        activity1_zt.append(activity1_full[:, zt_lo_min:zt_hi_min])
        exp_num.append(np.full(this_mins_awake.shape, ri))
        sleep_runs_all.extend(group_dat.get("sleep_runs", [[]] * n_flies_this_exp)[:n_flies_this_exp])
        active_runs_all.extend(group_dat.get("active_runs", [[]] * n_flies_this_exp)[:n_flies_this_exp])

    if not sleep_binary:
        return None

    return {
        "activity_per_min": np.vstack(activity_per_min),
        "sleep_per_min": np.vstack(sleep_per_min),
        "mins_awake": np.concatenate(mins_awake),
        "sleep_binary": np.vstack(sleep_binary),
        "activity1_zt": np.vstack(activity1_zt),
        "exp_num": np.concatenate(exp_num),
        "sleep_runs": sleep_runs_all,
        "active_runs": active_runs_all,
        "group_name": group_name_2_match,
        "win_lo": win_lo,
        "win_hi": win_hi,
    }


def bouts_starting_in_window(runs, win_lo, win_hi):
    """runs: list of (start, end) GLOBAL 0-based inclusive minute indices (spanning
    the fly's whole multi-day recording), sorted by start. Returns (start, clipped_end)
    for bouts whose START falls within [win_lo, win_hi) -- bouts already in progress
    when the window opens are excluded (this is what fixes the day/window-boundary
    'instant latency, phantom bout' artifact), while bouts that start inside the
    window are kept, with their length clipped to the window's own end."""
    out = []
    for start, end in runs:
        if win_lo <= start < win_hi:
            out.append((start, min(end, win_hi - 1)))
    return out


def summarize_windowed_bouts(runs, win_lo, win_hi):
    """(num_bouts, mean_len, median_len, longest_len, latency) for bouts starting
    within [win_lo, win_hi). Latency is 1-based minutes from win_lo to the first
    such bout; all-zeros if none start in the window (e.g. asleep the whole time)."""
    qualifying = bouts_starting_in_window(runs, win_lo, win_hi)
    if not qualifying:
        return 0.0, 0.0, 0.0, 0.0, 0.0
    lengths = [e - s + 1 for s, e in qualifying]
    latency = qualifying[0][0] - win_lo + 1
    return float(len(lengths)), float(np.mean(lengths)), float(np.median(lengths)), float(max(lengths)), float(latency)


def _ws(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def _write_col(ws, row_start, col, values):
    for i, v in enumerate(values):
        ws.cell(row=row_start + i, column=col, value=(None if (isinstance(v, float) and math.isnan(v)) else float(v)))


def _write_row(ws, row, col_start, values):
    for i, v in enumerate(values):
        ws.cell(row=row, column=col_start + i, value=float(v))


def _write_matrix(ws, row_start, col_start, matrix):
    for i, r in enumerate(matrix):
        for j, v in enumerate(r):
            ws.cell(row=row_start + i, column=col_start + j, value=(None if math.isnan(v) else float(v)))


def read_mat2read_txt(path):
    """mat2read as plain text: 'rootdir,matname,group1,group2,...' per line, no
    header. '#' starts a comment; blank lines are ignored."""
    rows = []
    with open(path, "r") as f:
        for line in f:
            line = line.split("#", 1)[0].rstrip("\n")
            if not line.strip():
                continue
            parts = [p.strip() for p in line.split(",")]
            if len(parts) < 3:
                continue
            rows.append(tuple(parts))
    return rows


# One "value per fly per day" sheet each; order here fixes both the sheet-creation
# order and the value order zipped together in step2_build_graphpad_export below.
SIMPLE_METRIC_SHEETS = [
    "Sleep (mins)",
    "Activity Counts Per min",
    "Total Activity Counts",
    "Percent Time Rest",
    "Percent Time Active",
    "Mean sleep bout w ends (mins)",
    "Median sleep bout (mins)",
    "Num sleep bouts",
    "Longest sleep bout (min)",
    "Time to first sleep bout (min)",
    "Num active bouts",
    "Mean active bout length (mins)",
    "Median active bout len (mins)",
    "Mean counts per activity bout",
    "Peak activity (counts per min)",
]


def step2_build_graphpad_export(state):
    print("\n--- Step 2: Aggregate experiments into a GraphPad Excel export ---")
    default_folder = state.get("last_folder")
    xl2read_path = None
    if default_folder:
        default_folder = Path(default_folder)
        for candidate in (f"{default_folder.name}_mat2read.txt", f"{default_folder.name}_mat2read.xlsx"):
            if (default_folder / candidate).is_file():
                xl2read_path = default_folder / candidate
                break
    if xl2read_path:
        print(f"Using {xl2read_path} (found automatically)")
    else:
        xl2read_path = ask_file(
            'Select the "mat2read" file listing the experiments to aggregate (.xlsx or .txt)',
            default_folder,
            filetypes=(("mat2read files", "*.xlsx *.txt"), ("Excel files", "*.xlsx"), ("Text files", "*.txt")),
        )
    primedir = xl2read_path.parent
    state["last_folder"] = primedir

    default_max_days = state.get("max_days", 3)
    max_days = ask_number("How many days should be aggregated?", default_max_days)
    state["max_days"] = max_days

    zt_windows = ask_zt_windows()

    if xl2read_path.suffix.lower() == ".txt":
        raw_rows = read_mat2read_txt(xl2read_path)
    else:
        wb_in = openpyxl.load_workbook(xl2read_path, data_only=True)
        ws_in = wb_in[wb_in.sheetnames[0]]
        raw_rows = [r for r in ws_in.iter_rows(values_only=True) if r and r[0] is not None]
    num_groups = len(raw_rows[0]) - 2

    all_outputs = []
    for zt_lo, zt_hi in zt_windows:
        if zt_lo == 0 and zt_hi == 24:
            out_name = re.sub(r"mat2read\.(?:xlsx|txt)$", "_24hrs_multiColumnByFly.xlsx", xl2read_path.name)
        else:
            out_name = re.sub(
                r"mat2read\.(?:xlsx|txt)$",
                f"_ZT{fmt_zt_hour(zt_lo)}to{fmt_zt_hour(zt_hi)}_multiColumn.xlsx",
                xl2read_path.name,
            )
        output_path = primedir / out_name

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for name in SIMPLE_METRIC_SHEETS:
            _write_row(_ws(wb, name), 1, 1, [0])  # placeholder to force sheet order; overwritten below
        for ws in wb.worksheets:
            ws.delete_rows(1)

        prev_row = 2
        for gi in range(num_groups):
            labels2write = []
            for di in range(1, max_days + 1):
                gathered = gather_group_day_data(raw_rows, gi, di, zt_lo, zt_hi)
                if gathered is None:
                    print(f"No data found for group column {gi+1}, day {di} - skipping.")
                    continue

                activity = gathered["activity_per_min"]
                sleep_per_min = gathered["sleep_per_min"]
                mins_awake = gathered["mins_awake"]
                sleep_binary = gathered["sleep_binary"]
                activity1_zt = gathered["activity1_zt"]
                exp_num = gathered["exp_num"]
                group_name = gathered["group_name"]
                sleep_runs = gathered["sleep_runs"]
                active_runs = gathered["active_runs"]
                win_lo, win_hi = gathered["win_lo"], gathered["win_hi"]
                window_len = win_hi - win_lo
                n_flies = sleep_binary.shape[0]

                total_sleep = np.nansum(sleep_binary, axis=1)
                total_activity_counts = np.nansum(activity, axis=1)
                total_activity_rate = total_activity_counts / mins_awake
                percent_rest = total_sleep / window_len * 100
                percent_active = mins_awake / window_len * 100

                num_sleep = np.zeros(n_flies)
                mean_sleep = np.zeros(n_flies)
                median_sleep = np.zeros(n_flies)
                longest_sleep = np.zeros(n_flies)
                latency = np.zeros(n_flies)
                num_active = np.zeros(n_flies)
                mean_active = np.zeros(n_flies)
                median_active = np.zeros(n_flies)
                mean_counts_bout = np.zeros(n_flies)
                peak_activity = np.zeros(n_flies)
                for xi in range(n_flies):
                    num_sleep[xi], mean_sleep[xi], median_sleep[xi], longest_sleep[xi], latency[xi] = \
                        summarize_windowed_bouts(sleep_runs[xi], win_lo, win_hi)
                    num_active[xi], mean_active[xi], median_active[xi], _longest_active, _latency_active = \
                        summarize_windowed_bouts(active_runs[xi], win_lo, win_hi)

                    qualifying_active = bouts_starting_in_window(active_runs[xi], win_lo, win_hi)
                    if qualifying_active:
                        totals = [np.nansum(activity1_zt[xi, s - win_lo:e - win_lo + 1]) for s, e in qualifying_active]
                        mean_counts_bout[xi] = float(np.mean(totals))
                    if activity1_zt.shape[1] > 0:
                        peak_activity[xi] = float(np.nanmax(activity1_zt[xi]))

                col_for_day = 2 + di  # A=1,B=2 -> di=1 => column C = 3
                sheet_name_day = f"Day {di} 30 min binned sleep"

                if prev_row == 2:
                    for name in SIMPLE_METRIC_SHEETS:
                        ws = _ws(wb, name)
                        ws.cell(row=1, column=1, value="Genotype")
                        ws.cell(row=1, column=2, value="Exp Row#")
                        ws.cell(row=1, column=3, value="Day 1")

                labels2write = [group_name] * n_flies

                values_by_sheet = dict(zip(SIMPLE_METRIC_SHEETS, [
                    total_sleep, total_activity_rate, total_activity_counts, percent_rest, percent_active,
                    mean_sleep, median_sleep, num_sleep, longest_sleep, latency,
                    num_active, mean_active, median_active, mean_counts_bout, peak_activity,
                ]))
                for name, values in values_by_sheet.items():
                    ws_stat = _ws(wb, name)
                    _write_col(ws_stat, prev_row, col_for_day, values)
                    if di == 1:
                        for i, lbl in enumerate(labels2write):
                            ws_stat.cell(row=prev_row + i, column=1, value=lbl)
                        _write_col(ws_stat, prev_row, 2, exp_num)

                ws_day = _ws(wb, sheet_name_day)
                if prev_row == 2:
                    ws_day.cell(row=1, column=1, value="Genotype")
                    ws_day.cell(row=1, column=2, value="Exp Row#")
                    ws_day.cell(row=1, column=3, value=f"Day {di}")
                for i, lbl in enumerate(labels2write):
                    ws_day.cell(row=prev_row + 1 + i, column=1, value=lbl)
                _write_col(ws_day, prev_row + 1, 2, exp_num)
                _write_matrix(ws_day, prev_row + 1, 3, sleep_per_min)
                if di == 1 and gi == 0:
                    _write_row(ws_day, prev_row, 3, np.arange(0.5, 24.5, 0.5))

                print(f'Wrote day {di} for "{group_name}" ({n_flies} flies) to column {col_for_day}')
            prev_row += len(labels2write)

        wb.save(output_path)
        print(f"Saved {output_path}")
        all_outputs.append(output_path)

    state["output_xlsx"] = all_outputs
    return state


# =============================================================================
# Step 3: remove dead/incomplete animals
# =============================================================================
def _remove_dead_animals_from_file(input_path):
    """Clean one multi-column export; returns the path to the _edited file."""
    output_path = input_path.with_name(input_path.stem + "_edited" + input_path.suffix)

    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    ws_act = wb_in["Activity Counts Per min"]
    act_rows = list(ws_act.iter_rows(values_only=True))
    dead_flies = set()
    for i, row in enumerate(act_rows[1:], start=0):  # i = 0-based fly index
        values = row[1:]  # skip label column (A)
        if any(v is None or (isinstance(v, float) and math.isnan(v)) for v in values):
            dead_flies.add(i + 2)  # matches MATLAB's find(...)+1 alignment

    day_sheet_names = [n for n in wb_in.sheetnames if n.startswith("Day ")]
    other_sheet_names = [n for n in wb_in.sheetnames if n not in day_sheet_names]

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for name in day_sheet_names:
        ws_in = wb_in[name]
        rows = list(ws_in.iter_rows(values_only=True))
        # Day sheets carry one extra metadata row (the time axis) before the header
        # row is stripped, so the row to drop is offset by one relative to dead_flies.
        rows_to_drop = {r + 1 for r in dead_flies}
        kept_rows = [r for idx, r in enumerate(rows, start=1) if idx not in rows_to_drop]
        ws_out = wb_out.create_sheet(name)
        for row in kept_rows:
            ws_out.append(row)

    for name in other_sheet_names:
        ws_in = wb_in[name]
        rows = list(ws_in.iter_rows(values_only=True))
        kept_rows = [r for idx, r in enumerate(rows, start=1) if idx not in dead_flies]
        ws_out = wb_out.create_sheet(name)
        for row in kept_rows:
            ws_out.append(row)

    ws_dead = wb_out.create_sheet("Dead Flies")
    for i, r in enumerate(sorted(dead_flies), start=1):
        ws_dead.cell(row=i, column=1, value=r)

    wb_out.save(output_path)
    print(f"Removed {len(dead_flies)} dead/incomplete animal row(s) from {input_path.name}. Saved {output_path}")
    return output_path


def step3_remove_dead_animals(state):
    """Returns the updated state, with output_xlsx pointed at the _edited file(s)
    so Step 4 defaults to visualizing the cleaned data."""
    print("\n--- Step 3: Remove dead animals from a GraphPad Excel export ---")
    prev_outputs = state.get("output_xlsx")
    if prev_outputs:
        names = ", ".join(p.name for p in prev_outputs)
        use_last = ask_choice(
            f"Clean the file(s) just created in step 2 ({names})?",
            ["Yes", "No, let me pick a different file"],
            0,
        )
    else:
        use_last = "No, let me pick a different file"

    if use_last == "Yes":
        input_paths = prev_outputs
    else:
        input_paths = [ask_file("Select the multi-column Excel file to clean", state.get("last_folder"))]

    state["output_xlsx"] = [_remove_dead_animals_from_file(p) for p in input_paths]
    return state


# =============================================================================
# Step 4: Prism-ready plots & tables
# =============================================================================
def step4_generate_prism_export(state):
    print("\n--- Step 4: Generate Prism-ready plots & tables ---")
    prev_outputs = state.get("output_xlsx")
    if prev_outputs:
        names = ", ".join(p.name for p in prev_outputs)
        use_last = ask_choice(
            f"Use the multi-column export(s) just created ({names})?",
            ["Yes", "No, let me pick a different file"],
            0,
        )
    else:
        use_last = "No, let me pick a different file"

    if use_last == "Yes":
        input_paths = prev_outputs
    else:
        input_paths = [ask_file("Select the multi-column Excel export to visualize", state.get("last_folder"))]

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import sleep_pipeline

    for input_path in input_paths:
        out_dir = sleep_pipeline.run(str(input_path))
        print(f"Saved Prism-ready plots and data to {out_dir}")
    return state


# =============================================================================
# Main
# =============================================================================
def prompt_steps():
    labels = [
        "0) Import raw Monitor*.txt dumps -> Raw Data folder + channelList "
        "(optional -- only if starting from raw monitor dumps rather than an existing Raw Data folder)",
        "1) Process one experiment (channelList -> .pkl + PDF)",
        "2) Aggregate experiments into a GraphPad Excel export",
        "3) Remove dead animals from a GraphPad Excel export",
        "4) Generate Prism-ready plots & tables",
    ]
    if HAS_GUI:
        txt = simpledialog.askstring(
            "Input",
            "Which step(s) do you want to run? e.g. 1 2 3 4 (add 0 first if starting from raw Monitor*.txt dumps)\n"
            + "\n".join(labels),
            initialvalue="1 2 3 4",
        )
    else:
        print("\n".join(labels))
        txt = input('Enter step numbers to run, e.g. "1 2 3 4" [1 2 3 4]: ').strip()
    if not txt or not txt.strip():
        return {1, 2, 3, 4}
    return {int(x) for x in re.findall(r"\d+", txt)}


def main():
    print("\n=== DAM Sleep Analysis Pipeline ===")
    steps = prompt_steps()
    state = {"max_days": None, "output_xlsx": None, "last_folder": Path.cwd()}

    if 0 in steps:
        state = step0_import_monitor_data(state)
    if 1 in steps:
        state = step1_process_experiment(state)
    if 2 in steps:
        state = step2_build_graphpad_export(state)
    if 3 in steps:
        state = step3_remove_dead_animals(state)
    if 4 in steps:
        state = step4_generate_prism_export(state)

    print("\nPipeline finished.")


if __name__ == "__main__":
    main()
