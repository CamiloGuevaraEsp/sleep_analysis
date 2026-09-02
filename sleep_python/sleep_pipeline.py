#!/usr/bin/env python3
"""
Sleep behavior pipeline: DAM-style multi-column ZT0-24 sleep workbook -> plots + Prism-ready data.

Usage:
    python3 sleep_pipeline.py /path/to/20260804__ZT0to24_multiColumn.xlsx

Expects the workbook layout used by this lab's sleep export:
  - Some sheets are "binned sleep" time-courses, one per recording day, with a sheet
    name containing "binned" and "Day <N>" (e.g. "Day 1 30 min binned sleep1"). Columns:
    Genotype, Exp Row#, then one column per time bin across ZT0-24.
  - Other sheets are per-day summary metrics (e.g. "Sleep (mins)", "Num sleep bouts").
    Columns: Genotype, Exp Row#, then one column per recording day (Day 1, Day 2, ...).
  - Genotype names optionally end in "_male" / "_female" -- if ALL genotypes follow that
    convention, sex-split scatter plots (mean +/- SD) are generated in addition to the
    combined (all-groups) bar+dot plots (mean +/- SEM).

Output: a "<stem>_prism_export" folder next to the input file, containing:
  combined/         one bar+dot (mean +/- SEM) plot per metric, all genotype groups
  by_sex/           (only if sex convention detected) scatter (mean +/- SD) plots per
                    metric per sex, comparing genotypes within each sex; +sleep profile
  Prism_ready_data.xlsx   copy/paste-ready tables for every metric + the ZT sleep profile
"""
import sys
import os
import re
import numpy as np
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import font_manager

PALETTE = ['#2a78d6', '#eb6834', '#1baf7a', '#eda100', '#e87ba4', '#008300', '#4a3aa7', '#e34948']

plt.rcParams['font.family'] = 'Arial' if any('Arial' in f.name for f in font_manager.fontManager.ttflist) else 'DejaVu Sans'
plt.rcParams['axes.spines.top'] = False
plt.rcParams['axes.spines.right'] = False
plt.rcParams['axes.edgecolor'] = '#52514e'
plt.rcParams['axes.labelcolor'] = '#0b0b0b'
plt.rcParams['text.color'] = '#0b0b0b'
plt.rcParams['xtick.color'] = '#0b0b0b'
plt.rcParams['ytick.color'] = '#0b0b0b'
plt.rcParams['font.size'] = 11

RNG = np.random.default_rng(42)

# A DAM daily profile is one 24 h ZT cycle by construction. Held explicitly rather
# than inferred from max(timepoints), because the sliding-window probability traces
# are labelled by window CENTRE and so stop just short of 24 -- inferring the span
# from them would misplace the lights-off shading and the double-plot offset.
ZT_DAY_HOURS = 24.0

# Checked in this order, so 'P(Wake)'/'P(Doze)' win over the generic words.
FAMILY_TOKENS = ('P(Wake)', 'P(Doze)', 'activity', 'sleep')


# ---------------------------------------------------------------------------
# 1. Load & classify sheets
# ---------------------------------------------------------------------------

def load_workbook_sheets(path):
    """Returns (workbook, binned_sheets, metric_sheets).

    binned_sheets is a list of (family, day_num, sheet_name). A day-trace sheet is
    one whose name carries a day number and the word "binned" or "sliding"; its
    family is whichever known trace name it mentions, so each kind of trace becomes
    its own ZT profile instead of all of them being averaged into one. Sheets from
    older exports named just "Day N 30 min binned" fall back to 'sleep'.

    Families do not have to share a time grid -- the probability traces use a
    sliding window with a finer step than the 30-min sleep/activity bins -- so each
    sheet's own time-axis row is what gets used."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    binned_sheets = []   # (family, day_num, sheet_name)
    metric_sheets = []   # sheet_name
    for name in wb.sheetnames:
        ws = wb[name]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header or not header[0]:
            continue
        col0 = str(header[0]).strip().lower()
        if col0 != 'genotype':
            continue
        m = re.search(r'day\s*(\d+)', name, re.IGNORECASE)
        if m and re.search(r'binned|sliding', name, re.IGNORECASE):
            low = name.lower()
            family = next((t for t in FAMILY_TOKENS if t.lower() in low), 'sleep')
            binned_sheets.append((family, int(m.group(1)), name))
        else:
            metric_sheets.append(name)
    binned_sheets.sort(key=lambda t: (t[0], t[1]))
    return wb, binned_sheets, metric_sheets


# How each binned family is labelled and drawn. Colours follow Wiggin et al. 2020
# Fig. 1B: activity black, sleep blue, P(Doze) green, P(Wake) red.
PROFILE_STYLE = {
    'activity': {'label': 'Activity (counts/bin)', 'color': '#0b0b0b', 'order': 0},
    'sleep':    {'label': 'Sleep per bin (min)',   'color': '#2a78d6', 'order': 1},
    'P(Doze)':  {'label': 'P(Doze)',               'color': '#1baf7a', 'order': 2},
    'P(Wake)':  {'label': 'P(Wake)',               'color': '#e34948', 'order': 3},
}


def find_data_start_row(ws):
    for row in ws.iter_rows(min_row=2, values_only=False):
        v0 = row[0].value
        if v0 is not None and str(v0).strip().lower() != 'genotype':
            return row[0].row
    return 2


def read_metric_sheet(ws):
    """Returns geno_sequence (list, row order) and values (n_rows x n_days array)."""
    start_row = find_data_start_row(ws)
    max_col = ws.max_column
    n_days = max_col - 2
    genos, values = [], []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if row[0] is None:
            continue
        genos.append(row[0])
        values.append([row[2 + i] for i in range(n_days)])
    return genos, np.array(values, dtype=float), n_days


def read_binned_sheet(ws):
    start_row = find_data_start_row(ws)
    max_col = ws.max_column
    n_bins = max_col - 2
    # look for a numeric header row (time bins) somewhere before start_row
    timepoints = None
    for row in ws.iter_rows(min_row=1, max_row=start_row - 1, values_only=True):
        vals = row[2:2 + n_bins]
        if all(v is not None for v in vals) and all(isinstance(v, (int, float)) for v in vals):
            timepoints = [float(v) for v in vals]
            break
    if timepoints is None:
        bin_width = 24.0 / n_bins
        timepoints = [round(bin_width * (i + 1), 4) for i in range(n_bins)]
    genos, values = [], []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if row[0] is None:
            continue
        genos.append(row[0])
        values.append([row[2 + i] for i in range(n_bins)])
    return genos, np.array(values, dtype=float), timepoints


# ---------------------------------------------------------------------------
# 2. Aggregate
# ---------------------------------------------------------------------------

def detect_groups_and_sex(genotypes):
    order = list(dict.fromkeys(genotypes))  # first-appearance order, de-duplicated
    has_sex = all(re.search(r'_(male|female)$', g, re.IGNORECASE) for g in order)
    base_order = None
    sex_of = {}
    if has_sex:
        bases = []
        for g in order:
            m = re.search(r'^(.*)_(male|female)$', g, re.IGNORECASE)
            base, sex = m.group(1), m.group(2).lower()
            sex_of[g] = sex
            if base not in bases:
                bases.append(base)
        base_order = bases
    return order, has_sex, sex_of, base_order


def build_dataset(path):
    wb, binned_sheets, metric_sheet_names = load_workbook_sheets(path)
    if not metric_sheet_names and not binned_sheets:
        raise ValueError("No sheets with a 'Genotype' header column were found -- "
                          "this file doesn't match the expected layout.")

    metrics = {}  # metric_name -> dict(genos=[...], values=n x days array, n_days=int)
    for name in metric_sheet_names:
        genos, values, n_days = read_metric_sheet(wb[name])
        metrics[name] = {'genos': genos, 'values': values, 'n_days': n_days}

    families = {}
    for family, day_num, name in binned_sheets:
        families.setdefault(family, []).append((day_num, name))

    profiles = {}
    for family, sheets in families.items():
        mats, day_nums, genos_ref, timepoints = [], [], None, None
        for day_num, name in sorted(sheets):
            genos, values, tp = read_binned_sheet(wb[name])
            if genos_ref is None:
                genos_ref = genos
                timepoints = tp
            elif genos != genos_ref:
                print(f"  WARNING: fly order in '{name}' differs from the first binned "
                      f"sheet -- averaging by position anyway.", file=sys.stderr)
            mats.append(values)
            day_nums.append(day_num)
        min_rows = min(m.shape[0] for m in mats)
        mats = [m[:min_rows] for m in mats]
        # One profile per fly = its average across recording days (Wiggin et al.
        # 2020 average each individual's profile over days before pooling).
        avg = np.nanmean(np.stack(mats, axis=0), axis=0)

        # Continuous multi-day timeline: each day's bins laid end-to-end (day N
        # starts at (N-1) * day_span hours) rather than averaged into one 24h day.
        day_span = max(timepoints)
        continuous_timepoints = [day_idx * day_span + t for day_idx in range(len(mats)) for t in timepoints]
        continuous_values = np.hstack(mats)

        # If these came from a sliding window, remember its width so the plots can
        # say so -- a smoothed trace must never be shown as if it were raw bins.
        win = re.search(r'(\d+)\s*min\s*sliding', sorted(sheets)[0][1], re.IGNORECASE)
        profiles[family] = {
            'family': family,
            'window_minutes': int(win.group(1)) if win else None,
            'genos': genos_ref[:min_rows],
            'values': avg,
            'timepoints': timepoints,
            'day_numbers': day_nums,
            'day_span': day_span,
            'continuous_timepoints': continuous_timepoints,
            'continuous_values': continuous_values,
        }

    # 'profile' (singular) stays pointed at the sleep trace so every existing
    # caller and plot keeps working unchanged.
    profile = profiles.get('sleep')

    ref_genos = None
    for m in metrics.values():
        ref_genos = m['genos']
        break
    if ref_genos is None and profile is not None:
        ref_genos = profile['genos']
    group_order, has_sex, sex_of, base_order = detect_groups_and_sex(ref_genos)

    return {
        'metrics': metrics,
        'profile': profile,
        'profiles': profiles,
        'group_order': group_order,
        'has_sex': has_sex,
        'sex_of': sex_of,
        'base_order': base_order,
    }


def by_group(genos, values, group_order):
    out = {g: [] for g in group_order}
    for geno, row in zip(genos, values):
        out[geno].append(row)
    return {g: np.array(v) for g, v in out.items()}


# ---------------------------------------------------------------------------
# 3. Plots
# ---------------------------------------------------------------------------

def safe_filename(name):
    return re.sub(r'[^A-Za-z0-9_.-]+', '_', name).strip('_')


def plot_combined_bar_sem(metric_name, values_by_group, group_order, colors, out_dir):
    fig, ax = plt.subplots(figsize=(max(6.5, 1.35 * len(group_order)), 5.5))
    for xi, geno in enumerate(group_order):
        vals = values_by_group[geno]
        vals = vals[~np.isnan(vals)]
        mean = np.mean(vals)
        sem = np.std(vals, ddof=1) / np.sqrt(len(vals))
        color = colors[geno]
        ax.bar(xi, mean, width=0.62, color=color, alpha=0.25, edgecolor=color, linewidth=1.5, zorder=1)
        ax.errorbar(xi, mean, yerr=sem, fmt='none', ecolor='#0b0b0b', elinewidth=1.5, capsize=4, zorder=3)
        jitter = RNG.uniform(-0.16, 0.16, size=len(vals))
        ax.scatter(xi + jitter, vals, s=22, color=color, edgecolor='#0b0b0b', linewidth=0.4, zorder=2, alpha=0.9)
    ax.set_xticks(range(len(group_order)))
    ax.set_xticklabels(group_order, fontsize=8.5, rotation=25, ha='right')
    ax.set_xlim(-0.6, len(group_order) - 0.4)
    ax.set_ylabel(metric_name)
    ax.set_ylim(bottom=0)
    ax.set_title(f"{metric_name} (mean ± SEM)", fontsize=12, pad=10)
    fig.tight_layout()
    fname = os.path.join(out_dir, f"{safe_filename(metric_name)}.png")
    fig.savefig(fname, dpi=200)
    plt.close(fig)
    return fname


def plot_sex_scatter_sd(metric_name, sex_label, values_by_base, base_order, colors, out_dir):
    fig, ax = plt.subplots(figsize=(max(4.5, 1.6 * len(base_order)), 5.5))
    for xi, base in enumerate(base_order):
        vals = values_by_base[base]
        vals = vals[~np.isnan(vals)]
        mean = np.mean(vals)
        sd = np.std(vals, ddof=1)
        color = colors[base]
        jitter = RNG.uniform(-0.12, 0.12, size=len(vals))
        ax.scatter(xi + jitter, vals, s=26, color=color, edgecolor='#0b0b0b', linewidth=0.4, zorder=2, alpha=0.85)
        ax.errorbar(xi, mean, yerr=sd, fmt='none', ecolor='#0b0b0b', elinewidth=1.4, capsize=5, capthick=1.4, zorder=3)
        ax.hlines(mean, xi - 0.18, xi + 0.18, colors='#0b0b0b', linewidth=2.2, zorder=4)
    ax.set_xticks(range(len(base_order)))
    ax.set_xticklabels(base_order, fontsize=9.5, rotation=25, ha='right')
    ax.set_xlim(-0.6, len(base_order) - 0.4)
    ax.set_ylabel(metric_name)
    ax.set_ylim(bottom=0)
    ax.set_title(f"{metric_name}\n{sex_label.capitalize()}s (mean ± SD)", fontsize=12, pad=10)
    fig.tight_layout()
    fname = os.path.join(out_dir, f"{safe_filename(metric_name)}_{sex_label}.png")
    fig.savefig(fname, dpi=200)
    plt.close(fig)
    return fname


def plot_profile_combined(timepoints, values_by_group, group_order, colors, out_dir):
    fig, ax = plt.subplots(figsize=(9, 5.5))
    for geno in group_order:
        mat = values_by_group[geno]
        mean = np.nanmean(mat, axis=0)
        sem = np.nanstd(mat, axis=0, ddof=1) / np.sqrt(mat.shape[0])
        color = colors[geno]
        ax.plot(timepoints, mean, color=color, linewidth=2, label=geno)
        ax.fill_between(timepoints, mean - sem, mean + sem, color=color, alpha=0.18, linewidth=0)
    ax.set_xlabel('Zeitgeber Time (ZT, hours)')
    ax.set_ylabel('Sleep per bin (min)')
    ax.set_xlim(min(timepoints), max(timepoints))
    ax.set_title('Sleep profile (mean ± SEM)', fontsize=12, pad=10)
    ax.legend(loc='lower center', ncol=min(3, len(group_order)), fontsize=8, frameon=False,
              bbox_to_anchor=(0.5, -0.3 - 0.05 * (len(group_order) // 3)))
    fig.tight_layout()
    fname = os.path.join(out_dir, "Sleep_profile.png")
    fig.savefig(fname, dpi=200, bbox_inches='tight')
    plt.close(fig)
    return fname


def plot_profile_by_sex(timepoints, values_by_group, sex_of, base_order, colors, out_dir):
    fnames = []
    for sex in ['male', 'female']:
        genos_this_sex = [g for g, s in sex_of.items() if s == sex]
        if not genos_this_sex:
            continue
        fig, ax = plt.subplots(figsize=(8, 5.2))
        for base in base_order:
            geno = next((g for g in genos_this_sex if g.lower().startswith(base.lower())), None)
            if geno is None:
                continue
            mat = values_by_group[geno]
            mean = np.nanmean(mat, axis=0)
            sem = np.nanstd(mat, axis=0, ddof=1) / np.sqrt(mat.shape[0])
            color = colors[base]
            ax.plot(timepoints, mean, color=color, linewidth=2, label=base)
            ax.fill_between(timepoints, mean - sem, mean + sem, color=color, alpha=0.15, linewidth=0)
        ax.set_xlabel('Zeitgeber Time (ZT, hours)')
        ax.set_ylabel('Sleep per bin (min)')
        ax.set_xlim(min(timepoints), max(timepoints))
        ax.set_title(f'Sleep profile (mean ± SEM) -- {sex.capitalize()}s', fontsize=12, pad=10)
        ax.legend(loc='lower center', ncol=min(3, len(base_order)), fontsize=9, frameon=False, bbox_to_anchor=(0.5, -0.3))
        fig.tight_layout()
        fname = os.path.join(out_dir, f"Sleep_profile_{sex}.png")
        fig.savefig(fname, dpi=200, bbox_inches='tight')
        plt.close(fig)
        fnames.append(fname)
    return fnames


def _shade_nights(ax, n_days, day_span):
    """Subtle shading over the second half of each day (ZT12-24), the
    conventional lights-off window, to make day boundaries readable."""
    for d in range(n_days):
        ax.axvspan(d * day_span + day_span / 2, (d + 1) * day_span, color='#e8e6e1', alpha=0.35, zorder=0)
    for d in range(1, n_days):
        ax.axvline(d * day_span, color='#c9c7c2', linewidth=1, linestyle='--', zorder=0)


def plot_profile_continuous(timepoints, values_by_group, group_order, colors, day_span, n_days, out_dir):
    fig, ax = plt.subplots(figsize=(max(9, 3.2 * n_days), 5.5))
    _shade_nights(ax, n_days, day_span)
    for geno in group_order:
        mat = values_by_group[geno]
        mean = np.nanmean(mat, axis=0)
        sem = np.nanstd(mat, axis=0, ddof=1) / np.sqrt(mat.shape[0])
        color = colors[geno]
        ax.plot(timepoints, mean, color=color, linewidth=1.6, label=geno)
        ax.fill_between(timepoints, mean - sem, mean + sem, color=color, alpha=0.18, linewidth=0)
    ax.set_xlabel('Time (hours, continuous across recording days)')
    ax.set_ylabel('Sleep per bin (min)')
    ax.set_xlim(min(timepoints), max(timepoints))
    ax.set_title(f'Sleep profile across all {n_days} recording days (mean ± SEM)', fontsize=12, pad=10)
    ax.legend(loc='lower center', ncol=min(3, len(group_order)), fontsize=8, frameon=False,
              bbox_to_anchor=(0.5, -0.3 - 0.05 * (len(group_order) // 3)))
    fig.tight_layout()
    fname = os.path.join(out_dir, "Sleep_profile_continuous.png")
    fig.savefig(fname, dpi=200, bbox_inches='tight')
    plt.close(fig)
    return fname


def plot_profile_continuous_by_sex(timepoints, values_by_group, sex_of, base_order, colors, day_span, n_days, out_dir):
    fnames = []
    for sex in ['male', 'female']:
        genos_this_sex = [g for g, s in sex_of.items() if s == sex]
        if not genos_this_sex:
            continue
        fig, ax = plt.subplots(figsize=(max(8, 3 * n_days), 5.2))
        _shade_nights(ax, n_days, day_span)
        for base in base_order:
            geno = next((g for g in genos_this_sex if g.lower().startswith(base.lower())), None)
            if geno is None:
                continue
            mat = values_by_group[geno]
            mean = np.nanmean(mat, axis=0)
            sem = np.nanstd(mat, axis=0, ddof=1) / np.sqrt(mat.shape[0])
            color = colors[base]
            ax.plot(timepoints, mean, color=color, linewidth=1.6, label=base)
            ax.fill_between(timepoints, mean - sem, mean + sem, color=color, alpha=0.15, linewidth=0)
        ax.set_xlabel('Time (hours, continuous across recording days)')
        ax.set_ylabel('Sleep per bin (min)')
        ax.set_xlim(min(timepoints), max(timepoints))
        ax.set_title(f'Sleep profile across all {n_days} days (mean ± SEM) -- {sex.capitalize()}s', fontsize=12, pad=10)
        ax.legend(loc='lower center', ncol=min(3, len(base_order)), fontsize=9, frameon=False, bbox_to_anchor=(0.5, -0.3))
        fig.tight_layout()
        fname = os.path.join(out_dir, f"Sleep_profile_continuous_{sex}.png")
        fig.savefig(fname, dpi=200, bbox_inches='tight')
        plt.close(fig)
        fnames.append(fname)
    return fnames


def _mean_ci(mat):
    """Column-wise mean and half-width of the 95% CI of the mean, ignoring NaN.
    Uses 1.96 x SEM (normal approximation) rather than a t quantile, to avoid a
    scipy dependency -- with n >= 15 flies per group the difference is small, and
    every plot says so on the axis label. Columns with fewer than 2 usable values
    get a NaN interval instead of a fake zero-width one."""
    mean = np.nanmean(mat, axis=0)
    n = np.sum(~np.isnan(mat), axis=0)
    sd = np.nanstd(mat, axis=0, ddof=1)
    with np.errstate(invalid='ignore', divide='ignore'):
        half = 1.96 * sd / np.sqrt(n)
    half = np.where(n >= 2, half, np.nan)
    return mean, half


def _double_plot(timepoints, values, span=ZT_DAY_HOURS):
    """Chronobiology double-plot: repeat the 24 h cycle so ZT0-24 is shown twice,
    the second copy shifted by one day. Makes events near ZT0/ZT24 readable
    instead of being cut in half at the edges of the axis."""
    x = list(timepoints) + [t + span for t in timepoints]
    return np.array(x), np.concatenate([values, values])


def _shade_nights_zt(ax, span, n_cycles):
    for c in range(n_cycles):
        ax.axvspan(c * span + span / 2, (c + 1) * span, color='#d8d5cf', alpha=0.55,
                   zorder=0, linewidth=0)


def plot_wake_doze_profile(profiles, geno, group_order, out_dir, n_days):
    """Stacked double-plotted ZT profiles for one genotype -- Activity, % Time
    Asleep, P(Doze) and P(Wake) -- population mean with 95% CI band, lights-off
    shaded. Layout follows Wiggin et al. 2020 Fig. 1B."""
    available = [f for f in sorted(PROFILE_STYLE, key=lambda k: PROFILE_STYLE[k]['order'])
                 if f in profiles]
    if not available:
        return None

    fig, axes = plt.subplots(len(available), 1, figsize=(8.5, 2.05 * len(available)), sharex=True)
    axes = np.atleast_1d(axes)
    span = ZT_DAY_HOURS

    for ax, family in zip(axes, available):
        prof = profiles[family]
        mat = by_group(prof['genos'], prof['values'], group_order)[geno]
        if mat.ndim != 2 or mat.shape[0] == 0:
            plt.close(fig)
            return None
        values = mat
        label = PROFILE_STYLE[family]['label']
        # "% Time Asleep" is the sleep trace expressed as a percentage of the bin.
        if family == 'sleep':
            bin_minutes = ZT_DAY_HOURS * 60.0 / len(prof['timepoints'])
            values = mat / bin_minutes * 100.0
            label = '% Time Asleep'
        mean, half = _mean_ci(values)
        x, y = _double_plot(prof['timepoints'], mean)
        _, h = _double_plot(prof['timepoints'], half)
        _shade_nights_zt(ax, span, 2)
        color = PROFILE_STYLE[family]['color']
        ax.plot(x, y, color=color, linewidth=1.8, zorder=3)
        ax.fill_between(x, y - h, y + h, color=color, alpha=0.22, linewidth=0, zorder=2)
        if prof.get('window_minutes'):
            label += f"\n({prof['window_minutes']} min sliding)"
        ax.set_ylabel(label, fontsize=9.5)
        ax.set_xlim(0, 2 * span)
        if family.startswith('P('):
            ax.set_ylim(0, None)

    axes[-1].set_xlabel('Zeitgeber Time (ZT, hours) -- double plotted')
    axes[-1].set_xticks(np.arange(0, 2 * span + 1, 6))
    n = sum(1 for g in profiles[available[0]]['genos'] if g == geno)
    axes[0].set_title(f"{geno}  (n = {n} flies, mean of {n_days} day{'s' if n_days > 1 else ''})\n"
                      "shaded band = 95% CI of the mean (1.96 x SEM)", fontsize=10.5, pad=8)
    fig.tight_layout()
    fname = os.path.join(out_dir, f"Wake_doze_profile_{safe_filename(geno)}.png")
    fig.savefig(fname, dpi=200)
    plt.close(fig)
    return fname


def plot_profile_metric(family, profiles, group_order, colors, out_dir):
    """One ZT profile per binned family, all genotypes overlaid -- the comparison
    view that Fig. 1B (single genotype) doesn't give you."""
    prof = profiles[family]
    by_group_data = by_group(prof['genos'], prof['values'], group_order)
    span = ZT_DAY_HOURS
    fig, ax = plt.subplots(figsize=(9, 5.5))
    _shade_nights_zt(ax, span, 2)
    for geno in group_order:
        mean, half = _mean_ci(by_group_data[geno])
        x, y = _double_plot(prof['timepoints'], mean)
        _, h = _double_plot(prof['timepoints'], half)
        color = colors[geno]
        ax.plot(x, y, color=color, linewidth=1.8, label=geno, zorder=3)
        ax.fill_between(x, y - h, y + h, color=color, alpha=0.16, linewidth=0, zorder=2)
    ax.set_xlabel('Zeitgeber Time (ZT, hours) -- double plotted')
    ax.set_ylabel(PROFILE_STYLE[family]['label'])
    ax.set_xlim(0, 2 * span)
    ax.set_xticks(np.arange(0, 2 * span + 1, 6))
    if family.startswith('P('):
        ax.set_ylim(0, None)
    how = (f"{prof['window_minutes']} min sliding window"
           if prof.get('window_minutes') else f"{int(ZT_DAY_HOURS * 60 / len(prof['timepoints']))} min bins")
    ax.set_title(f"{PROFILE_STYLE[family]['label']} across the day ({how})\n"
                 "shaded band = 95% CI of the mean (1.96 x SEM)", fontsize=12, pad=10)
    ax.legend(loc='lower center', ncol=min(3, len(group_order)), fontsize=8, frameon=False,
              bbox_to_anchor=(0.5, -0.32 - 0.05 * (len(group_order) // 3)))
    fig.tight_layout()
    fname = os.path.join(out_dir, f"{safe_filename(family)}_ZT_profile.png")
    fig.savefig(fname, dpi=200, bbox_inches='tight')
    plt.close(fig)
    return fname


# ---------------------------------------------------------------------------
# 4. Prism-ready workbook
# ---------------------------------------------------------------------------

def write_column_table(ws, start_row, start_col, title, group_order, group_to_values):
    HEADER_FONT = Font(bold=True)
    TITLE_FONT = Font(bold=True, size=12)
    ws.cell(row=start_row, column=start_col, value=title).font = TITLE_FONT
    header_row = start_row + 1
    for j, geno in enumerate(group_order):
        ws.cell(row=header_row, column=start_col + j, value=geno).font = HEADER_FONT
    max_n = max(len(group_to_values[g]) for g in group_order)
    for j, geno in enumerate(group_order):
        for i, v in enumerate(group_to_values[geno]):
            if v is None or (isinstance(v, float) and np.isnan(v)):
                continue
            ws.cell(row=header_row + 1 + i, column=start_col + j, value=round(float(v), 3))
    return header_row + 1 + max_n


def _write_profile_sheet(wb, sheet_name, title, timepoints, by_group_data, group_order):
    """Write one Prism 'XY' table sheet (X = time, then Mean/SEM/N sub-columns
    per group) for a sleep profile -- used for both the single averaged-24h
    ZT profile and the continuous multi-day profile."""
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
    header_row, col = 2, 2
    ws.cell(row=header_row, column=1, value="Time").font = Font(bold=True)
    group_col_start = {}
    for geno in group_order:
        group_col_start[geno] = col
        ws.cell(row=header_row, column=col, value=geno).font = Font(bold=True)
        ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=col + 2)
        for k, label in enumerate(['Mean', 'SEM', 'N']):
            ws.cell(row=header_row + 1, column=col + k, value=label).font = Font(bold=True)
        col += 3
    data_start = header_row + 2
    for i, tp in enumerate(timepoints):
        r = data_start + i
        ws.cell(row=r, column=1, value=float(tp))
        for geno in group_order:
            vals = by_group_data[geno][:, i]
            vals = vals[~np.isnan(vals)]
            if len(vals) == 0:
                continue
            mean, sd = float(np.mean(vals)), float(np.std(vals, ddof=1))
            sem, n = sd / np.sqrt(len(vals)), int(len(vals))
            c0 = group_col_start[geno]
            ws.cell(row=r, column=c0, value=round(mean, 3))
            ws.cell(row=r, column=c0 + 1, value=round(sem, 3))
            ws.cell(row=r, column=c0 + 2, value=n)
    return ws


def write_prism_workbook(dataset, out_path):
    group_order = dataset['group_order']
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('README')
    ws.column_dimensions['A'].width = 100
    lines = [
        "How to use this workbook with GraphPad Prism",
        "",
        "Each metric has its own sheet, formatted as a Prism 'Column' table: one column per",
        "genotype group, individual fly values stacked below the header.",
        "",
        "To use: select the header row + data block, copy, create a new Prism 'Column' data",
        "table (individual replicate values), and paste into the top-left cell.",
        "",
        "'Average of all days' block: one value per fly, averaged across all recording days.",
        "'Day N' blocks: same layout, one recording day at a time.",
        "",
        "'ZT Sleep Profile' sheet is a Prism 'XY' table (X = time, then Mean/SEM/N",
        "sub-columns per group) -- paste using 'Enter/import data using format: Mean, SEM",
        "(or SD) and N'.",
        "",
        "'Continuous Sleep Profile' sheet (only present if there is more than one recording",
        "day) is the same XY layout, but X runs continuously across all recording days",
        "(day N starts at (N-1) x day-length hours) instead of averaging all days into one",
        "24h ZT profile -- matches the 'Sleep_profile_continuous.png' plot.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13)

    for metric_name, m in dataset['metrics'].items():
        by_group_data = by_group(m['genos'], m['values'], group_order)
        avg_by_group = {g: np.nanmean(by_group_data[g], axis=1) for g in group_order}
        sheet_title = re.sub(r'[\\/*?:\[\]]', '', metric_name)[:31]
        ws = wb.create_sheet(sheet_title)
        for j in range(len(group_order)):
            ws.column_dimensions[get_column_letter(j + 1)].width = 18
        next_row = write_column_table(ws, 1, 1, f"{metric_name} -- Average of all days", group_order, avg_by_group)
        next_row += 2
        for day_idx in range(m['n_days']):
            day_vals = {g: by_group_data[g][:, day_idx] for g in group_order}
            next_row = write_column_table(ws, next_row, 1, f"{metric_name} -- Day {day_idx + 1}", group_order, day_vals)
            next_row += 2

    profile = dataset['profile']
    if profile is not None:
        by_group_data = by_group(profile['genos'], profile['values'], group_order)
        _write_profile_sheet(wb, 'ZT Sleep Profile', "Sleep per bin, group Mean/SEM/N",
                              profile['timepoints'], by_group_data, group_order)

        n_days = len(profile['day_numbers'])
        if n_days > 1:
            by_group_data_cont = by_group(profile['genos'], profile['continuous_values'], group_order)
            _write_profile_sheet(wb, 'Continuous Sleep Profile',
                                  f"Sleep per bin across all {n_days} recording days, group Mean/SEM/N "
                                  "(X = continuous hours, days laid end-to-end)",
                                  profile['continuous_timepoints'], by_group_data_cont, group_order)

    # ZT profiles for the other binned traces (activity, P(Wake), P(Doze)) --
    # the numbers behind the Fig. 1B plots, in the same Prism XY layout.
    for family, prof in sorted(dataset.get('profiles', {}).items()):
        if family == 'sleep':
            continue  # already written above as 'ZT Sleep Profile'
        label = PROFILE_STYLE.get(family, {}).get('label', family)
        by_group_data = by_group(prof['genos'], prof['values'], group_order)
        sheet_title = re.sub(r'[\\/*?:\[\]]', '', f"{family} ZT Profile")[:31]
        _write_profile_sheet(wb, sheet_title, f"{label} per bin, group Mean/SEM/N",
                             prof['timepoints'], by_group_data, group_order)

    wb.save(out_path)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def run(path):
    """Run the full plotting + Prism-export pipeline for one workbook.
    Callable directly (e.g. from another pipeline script) or via main()."""
    if not os.path.isfile(path):
        raise FileNotFoundError(f"File not found: {path}")

    stem = os.path.splitext(os.path.basename(path))[0]
    out_dir = os.path.join(os.path.dirname(os.path.abspath(path)), f"{stem}_prism_export")
    combined_dir = os.path.join(out_dir, 'combined')
    os.makedirs(combined_dir, exist_ok=True)

    print(f"Loading {path} ...")
    dataset = build_dataset(path)
    group_order = dataset['group_order']
    print(f"Detected {len(group_order)} groups: {group_order}")
    print(f"Sex convention detected: {dataset['has_sex']}")
    print(f"Metric sheets: {list(dataset['metrics'].keys())}")
    print(f"Time-course profile present: {dataset['profile'] is not None}")

    colors = {g: PALETTE[i % len(PALETTE)] for i, g in enumerate(group_order)}

    for metric_name, m in dataset['metrics'].items():
        avg_vals = np.nanmean(m['values'], axis=1)
        values_by_group = by_group(m['genos'], avg_vals.reshape(-1, 1), group_order)
        values_by_group = {g: v[:, 0] for g, v in values_by_group.items()}
        fname = plot_combined_bar_sem(metric_name, values_by_group, group_order, colors, combined_dir)
        print("  Saved", fname)

    if dataset['profile'] is not None:
        prof = dataset['profile']
        values_by_group = by_group(prof['genos'], prof['values'], group_order)
        fname = plot_profile_combined(prof['timepoints'], values_by_group, group_order, colors, combined_dir)
        print("  Saved", fname)

        n_days = len(prof['day_numbers'])
        if n_days > 1:
            values_by_group_cont = by_group(prof['genos'], prof['continuous_values'], group_order)
            fname = plot_profile_continuous(prof['continuous_timepoints'], values_by_group_cont, group_order,
                                             colors, prof['day_span'], n_days, combined_dir)
            print("  Saved", fname)

    # One stacked, double-plotted figure per genotype (layout of Wiggin et al.
    # 2020 Fig. 1B), plus an all-genotypes overlay for each binned trace.
    profiles = dataset.get('profiles', {})
    extra_families = [f for f in profiles if f != 'sleep']
    if extra_families:
        profile_dir = os.path.join(out_dir, 'wake_doze_profiles')
        os.makedirs(profile_dir, exist_ok=True)
        n_days = len(profiles[next(iter(profiles))]['day_numbers'])
        for geno in group_order:
            fname = plot_wake_doze_profile(profiles, geno, group_order, profile_dir, n_days)
            if fname:
                print("  Saved", fname)
        for family in sorted(profiles, key=lambda k: PROFILE_STYLE.get(k, {}).get('order', 99)):
            if family == 'sleep':
                continue  # already covered by Sleep_profile.png
            fname = plot_profile_metric(family, profiles, group_order, colors, profile_dir)
            print("  Saved", fname)

    if dataset['has_sex']:
        by_sex_dir = os.path.join(out_dir, 'by_sex')
        os.makedirs(by_sex_dir, exist_ok=True)
        base_order = dataset['base_order']
        sex_of = dataset['sex_of']
        base_colors = {b: PALETTE[i % len(PALETTE)] for i, b in enumerate(base_order)}

        for metric_name, m in dataset['metrics'].items():
            avg_vals = np.nanmean(m['values'], axis=1)
            values_by_group = by_group(m['genos'], avg_vals.reshape(-1, 1), group_order)
            values_by_group = {g: v[:, 0] for g, v in values_by_group.items()}
            for sex in ['male', 'female']:
                values_by_base = {}
                for base in base_order:
                    geno = next((g for g in group_order if sex_of.get(g) == sex and g.lower().startswith(base.lower())), None)
                    if geno is None:
                        continue
                    values_by_base[base] = values_by_group[geno]
                if len(values_by_base) == len(base_order):
                    fname = plot_sex_scatter_sd(metric_name, sex, values_by_base, base_order, base_colors, by_sex_dir)
                    print("  Saved", fname)

        if dataset['profile'] is not None:
            prof = dataset['profile']
            values_by_group = by_group(prof['genos'], prof['values'], group_order)
            fnames = plot_profile_by_sex(prof['timepoints'], values_by_group, sex_of, base_order, base_colors, by_sex_dir)
            for fname in fnames:
                print("  Saved", fname)

            n_days = len(prof['day_numbers'])
            if n_days > 1:
                values_by_group_cont = by_group(prof['genos'], prof['continuous_values'], group_order)
                fnames = plot_profile_continuous_by_sex(prof['continuous_timepoints'], values_by_group_cont, sex_of,
                                                         base_order, base_colors, prof['day_span'], n_days, by_sex_dir)
                for fname in fnames:
                    print("  Saved", fname)

    prism_path = os.path.join(out_dir, 'Prism_ready_data.xlsx')
    write_prism_workbook(dataset, prism_path)
    print("  Saved", prism_path)

    print(f"\nDone. All output in: {out_dir}")
    return out_dir


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    try:
        run(sys.argv[1])
    except (FileNotFoundError, ValueError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
